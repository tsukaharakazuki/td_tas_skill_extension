"""
Excel → Treasure Data アップロードスクリプト（テンプレート）

使い方:
  1. 下記の設定値（EXCEL_PATH, DATABASE, TD_API_KEY, TD_ENDPOINT）を書き換える
  2. SHEET_CONFIG にシート名→テーブル名のマッピングを定義する
  3. COLUMN_MAPS に日本語カラム名→英語カラム名のマッピングを定義する
  4. python3 upload.py を実行する

APIキーの取得:
  TDコンソール（https://console.us01.treasuredata.com）
  → Settings → API Keys → Write権限のキーをコピー
"""
import openpyxl
import math
import os
import json
import subprocess
import tempfile
import time
import re
from datetime import datetime

# ── 設定 ─────────────────────────────────────────────────────────────────────

EXCEL_PATH = "path/to/file.xlsx"           # アップロードするExcelファイルのパス
DATABASE   = "target_database"             # アップロード先DB名
TD_API_KEY = "<account_id>/xxxxxxxxxx..."  # TDコンソールから取得したAPIキー
TD_ENDPOINT = "https://api.treasuredata.com"  # US01
# JP リージョン: https://api.treasuredata.co.jp
# EU リージョン: https://api.eu01.treasuredata.com

EXCEL_FILENAME = os.path.basename(EXCEL_PATH)
NOW_TS = int(time.time())  # 全レコード共通の time 値

# ── シート → テーブルのマッピング ─────────────────────────────────────────────
# 形式: "Excelシート名": ("tdテーブル名", ヘッダー行インデックス)
#   ヘッダー行インデックス = 0 → 1行目がヘッダー（通常）
#   ヘッダー行インデックス = 1 → 2行目がヘッダー（1行目がタイトル行の場合）

SHEET_CONFIG = {
    "シートA": ("table_a", 0),
    "シートB": ("table_b", 0),
    "シートC": ("table_c", 1),  # 2行目がヘッダーの場合
    # ...
}

# ── カラム名変換マップ（日本語→英語） ────────────────────────────────────────
# 未登録のカラムは normalize_col_name() で自動変換される
# （日本語が残る可能性があるため、なるべく全カラムを登録しておくことを推奨）

COLUMN_MAPS = {
    "table_a": {
        "日本語カラム名1": "english_col_name_1",
        "日本語カラム名2": "english_col_name_2",
        # ...
    },
    "table_b": {
        # ...
    },
}

# ─────────────────────────────────────────────────────────────────────────────


def clean_value(v):
    """NaN / Inf / Excelエラー値を None に変換"""
    if v is None:
        return None
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return None
    if isinstance(v, str) and v in ("#REF!", "#N/A", "#VALUE!", "#NAME?", "#DIV/0!"):
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    return v


def normalize_col_name(raw, col_map, index):
    """カラム名を正規化: マップ → 日付 → フォールバック"""
    if raw is None:
        return f"column_{index}"
    s = str(raw).strip()
    if s in col_map:
        return col_map[s]
    if isinstance(raw, datetime):
        return f"date_{raw.strftime('%Y_%m_%d')}"
    s_clean = re.sub(r'[^\w]', '_', s.lower())
    s_clean = re.sub(r'_+', '_', s_clean).strip('_')
    if not s_clean or s_clean[0].isdigit():
        s_clean = f"col_{s_clean}"
    return s_clean if s_clean else f"column_{index}"


def td(args):
    """td コマンド実行（APIキー・エンドポイント固定）"""
    return subprocess.run(
        ["td", "-k", TD_API_KEY, "-e", TD_ENDPOINT] + args,
        capture_output=True, text=True
    )


def ensure_table(table_name):
    """テーブルが存在しない場合のみ REST API で作成（DBリスト権限不要）"""
    r = subprocess.run(
        ["curl", "-s", "-H", f"Authorization: TD1 {TD_API_KEY}",
         f"{TD_ENDPOINT}/v3/table/show/{DATABASE}/{table_name}"],
        capture_output=True, text=True
    )
    if '"type"' in r.stdout:
        return  # 既存テーブル → そのまま使う
    # 存在しない場合は REST API で作成
    subprocess.run(
        ["curl", "-s", "-X", "POST",
         "-H", f"Authorization: TD1 {TD_API_KEY}",
         f"{TD_ENDPOINT}/v3/table/create/{DATABASE}/{table_name}/log"],
        capture_output=True, text=True
    )


def upload_table(table_name, records):
    """td table:import --json でアップロード"""
    ensure_table(table_name)

    with tempfile.NamedTemporaryFile(
        mode='w', suffix='.json', delete=False, encoding='utf-8'
    ) as f:
        tmp_path = f.name
        for rec in records:
            f.write(json.dumps({"time": NOW_TS, **rec}, ensure_ascii=False) + "\n")

    try:
        r = td(["table:import", DATABASE, table_name, "--json", "-t", "time", tmp_path])
        if r.returncode != 0:
            raise RuntimeError(f"table:import failed: {r.stderr[:300]}")
        for line in r.stdout.splitlines():
            if "imported" in line or "skipped" in line or "Error" in line:
                print(f"    {line.strip()}")
    finally:
        os.unlink(tmp_path)


def verify_counts(tables, wait_sec=30):
    """アップロード後の件数確認（REST API 経由）"""
    print(f"\n--- 件数確認（{wait_sec}秒待機後）---")
    time.sleep(wait_sec)
    for tbl in tables:
        r = subprocess.run(
            ["curl", "-s", "-H", f"Authorization: TD1 {TD_API_KEY}",
             f"{TD_ENDPOINT}/v3/table/show/{DATABASE}/{tbl}"],
            capture_output=True, text=True
        )
        try:
            count = json.loads(r.stdout).get("count", "?")
        except Exception:
            count = "?"
        print(f"  {tbl}: {count} rows")


# ── メイン処理 ────────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
uploaded_tables = []

for sheet_name, (table_name, header_row_idx) in SHEET_CONFIG.items():
    if sheet_name not in wb.sheetnames:
        print(f"  SKIP (not found): {sheet_name}")
        continue

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        print(f"  SKIP (empty): {sheet_name}")
        continue

    header_raw = all_rows[header_row_idx]
    data_rows  = all_rows[header_row_idx + 1:]

    col_map = COLUMN_MAPS.get(table_name, {})
    header  = [normalize_col_name(c, col_map, i) for i, c in enumerate(header_raw)]

    records = []
    for row in data_rows:
        if not any(v is not None for v in row):
            continue
        record = {header[i]: clean_value(v) for i, v in enumerate(row) if i < len(header)}
        record["excel_name"]       = EXCEL_FILENAME
        record["excel_sheet_name"] = sheet_name
        records.append(record)

    if not records:
        print(f"  SKIP (no data rows): {sheet_name} → {table_name}")
        continue

    print(f"\nUploading: [{sheet_name}] → {table_name} ({len(records)} rows, {len(header)} cols)")
    try:
        upload_table(table_name, records)
        print(f"  ✓ Done: {table_name}")
        uploaded_tables.append(table_name)
    except Exception as e:
        print(f"  ✗ Error: {table_name}: {e}")

verify_counts(uploaded_tables)
print("\nAll done.")
